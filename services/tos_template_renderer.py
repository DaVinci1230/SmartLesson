"""
Fixed Institutional Table of Specifications (TOS) Template Renderer
Matches exact institutional format from reference template.

Structure:
- Header: Metadata section (Name, Subject Code, Title, Semester, etc.)
- Fixed columns: Content | Outcomes | RBT | Hours | Bloom×6 | Items | Points
- Fixed rows: Always 15+ content rows minimum
- Footer: ROUNDINGS row, TOTAL row
- Formatting: Fixed widths, fixed heights, professional borders
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO


# ======================================================
# CONSTANTS - INSTITUTIONAL TEMPLATE STRUCTURE
# ======================================================

BLOOM_LEVELS = ["remember", "understand", "apply", "analyze", "evaluate", "create"]
BLOOM_LABELS = ["REMEMBERING", "UNDERSTANDING", "APPLYING", "ANALYZING", "EVALUATING", "CREATING"]

# Fixed dimensions matching institutional template
MIN_DATA_ROWS = 15  # Institutional template shows ~15 content rows
HEADER_START_ROW = 1
METADATA_ROWS = 5   # Name, Subject Code, Title, Total Items, Total Points
TABLE_HEADER_ROW = 7
DATA_START_ROW = 8

# Column configuration (explicit widths matching reference)
COLUMN_WIDTHS = {
    'A': 20,  # Course Content
    'B': 35,  # Learning Outcomes
    'C': 12,  # RBT
    'D': 12,  # No. of Hours
    # Bloom columns (E onwards): width 10 each for Items/Pts/ItemNo
}


# ======================================================
# SECTION 1: BUILD METADATA HEADER
# ======================================================

def render_institutional_header(ws, meta):
    """
    Render header with metadata exactly like institutional template.
    
    Layout:
    Row 1: [blank] [TITLE spanning across] [blank]
    Row 2: [blank]
    Row 3-7: LEFT side (A-B) | [blank] | RIGHT side (S-T)
    
    Metadata fields:
    - Name: Instructor name
    - Subject Code: Course code
    - Descriptive Title: Course title
    - Total Test Items: From total_items
    - Total Number of Points: From total_points (computed from question types)
    """
    
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # ===== ROW 1: TITLE =====
    ws.merge_cells("A1:O1")
    title = ws["A1"]
    title.value = "TABLE OF SPECIFICATIONS (TOS)"
    title.font = Font(bold=True, size=12)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20
    
    # ===== ROW 2: BLANK =====
    ws.row_dimensions[2].height = 5
    
    # ===== ROWS 3-7: METADATA (LEFT AND RIGHT) =====
    left_meta = [
        ("Name:", meta.get("name", "")),
        ("Subject Code:", meta.get("subject_code", "")),
        ("Descriptive Title:", meta.get("title", "")),
        ("Total Test Items:", meta.get("total_items", "")),
        ("Total Number of Points:", meta.get("total_points", "")),  # From question type distribution
    ]
    
    right_meta = [
        ("Semester:", meta.get("semester", "")),
        ("Class Schedule:", meta.get("schedule", "")),
        ("Course:", meta.get("course", "")),
        ("Exam Date:", meta.get("exam_date", "")),
    ]
    
    # LEFT SIDE metadata (Columns A-B)
    row = 3
    for label, value in left_meta:
        cell_label = ws[f"A{row}"]
        cell_label.value = label
        cell_label.font = Font(bold=True, size=10)
        cell_label.border = thin_border
        cell_label.alignment = Alignment(horizontal="left", vertical="center")
        
        cell_value = ws[f"B{row}"]
        cell_value.value = value
        cell_value.border = thin_border
        cell_value.alignment = Alignment(horizontal="left", vertical="center")
        
        ws.row_dimensions[row].height = 18
        row += 1
    
    # RIGHT SIDE metadata (Columns S-T)
    row = 3
    for label, value in right_meta:
        cell_label = ws[f"S{row}"]
        cell_label.value = label
        cell_label.font = Font(bold=True, size=10)
        cell_label.border = thin_border
        cell_label.alignment = Alignment(horizontal="left", vertical="center")
        
        cell_value = ws[f"T{row}"]
        cell_value.value = value
        cell_value.border = thin_border
        cell_value.alignment = Alignment(horizontal="left", vertical="center")
        
        row += 1


# ======================================================
# SECTION 2: CREATE TABLE HEADERS
# ======================================================

def create_table_header_row(ws, start_row=7):
    """
    Create table headers exactly matching institutional template.
    
    KEY FIX: Course Content and Learning Outcomes are NOT MERGED.
    They are separate, distinct columns with clear borders.
    
    Structure:
    Row 7 (primary): Course Content | Learning Outcomes | RBT | Hours | Bloom headers
    Row 8 (secondary): - | - | - | - | Items|Pts|Item# (for each Bloom)
    
    Returns: Column mapping dict
    """
    
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    header_font = Font(bold=True, size=9)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    col_map = {}
    
    # ===== PRIMARY HEADERS (Row 7) =====
    
    # Column A: Course Content (span 2 rows, NOT merged with Outcomes)
    cell = ws["A7"]
    cell.value = "Course\nContent"
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = header_alignment
    ws.merge_cells("A7:A8")
    
    # Column B: Learning Outcomes (span 2 rows, SEPARATE from Content)
    cell = ws["B7"]
    cell.value = "Learning\nOutcomes\n(From the Syllabus)"
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = header_alignment
    ws.merge_cells("B7:B8")
    
    # Column C: RBT (span 2 rows)
    cell = ws["C7"]
    cell.value = "RBT"
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = header_alignment
    ws.merge_cells("C7:C8")
    
    # Column D: No. of Hours Taught (span 2 rows)
    cell = ws["D7"]
    cell.value = "No. of\nHours\nTaught"
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = header_alignment
    ws.merge_cells("D7:D8")
    
    # ===== BLOOM HEADERS (Row 7, each spans 3 columns) =====
    col = 5  # Start at column E
    for bloom_label in BLOOM_LABELS:
        # Merge 3 columns for this Bloom level
        start_col = get_column_letter(col)
        end_col = get_column_letter(col + 2)
        merge_range = f"{start_col}7:{end_col}7"
        ws.merge_cells(merge_range)
        
        cell = ws[f"{start_col}7"]
        cell.value = bloom_label
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = header_alignment
        
        col_map[bloom_label.lower()] = col
        col += 3
    
    # ===== ITEMS & POINTS COLUMNS (Row 7, span 2 rows) =====
    items_col = get_column_letter(col)
    cell = ws[f"{items_col}7"]
    cell.value = "ITEMS"
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = header_alignment
    ws.merge_cells(f"{items_col}7:{items_col}8")
    col_map["items_col"] = col
    col += 1
    
    points_col = get_column_letter(col)
    cell = ws[f"{points_col}7"]
    cell.value = "POINTS"
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = header_alignment
    ws.merge_cells(f"{points_col}7:{points_col}8")
    col_map["points_col"] = col
    
    # ===== SECONDARY HEADERS (Row 8) - Bloom sub-headers =====
    col = 5
    for _ in BLOOM_LABELS:
        for sub_header in ["No. of\nItems", "Pts", "Item\nNo."]:
            cell = ws.cell(row=8, column=col)
            cell.value = sub_header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = header_alignment
            col += 1
    
    # Set header row heights
    ws.row_dimensions[7].height = 25
    ws.row_dimensions[8].height = 25
    
    return col_map


# ======================================================
# SECTION 3: INJECT DATA INTO FIXED CELLS
# ======================================================

def inject_data_rows(ws, col_map, outcomes, tos_matrix, start_row=9):
    """
    Inject outcome data into fixed grid, always filling MIN_DATA_ROWS.
    
    Args:
        ws: Worksheet
        col_map: Column mapping from create_table_header_row()
        outcomes: List of outcomes [{text, hours, bloom_level}, ...]
        tos_matrix: {bloom: {idx: count}, ...}
        start_row: Where data rows begin (default 9)
    """
    
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    data_alignment_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    data_alignment_center = Alignment(horizontal="center", vertical="center")
    
    # Normalize TOS matrix keys
    normalized_tos = {
        k.strip().lower(): v
        for k, v in tos_matrix.items()
    }
    
    row_totals = {}
    current_row = start_row
    
    # ===== FILL ALL ROWS (MIN_DATA_ROWS) =====
    for row_idx in range(MIN_DATA_ROWS):
        
        # Has data for this row?
        if row_idx < len(outcomes):
            outcome = outcomes[row_idx]
            
            # Column A: Course Content (empty - user fills)
            cell = ws.cell(row=current_row, column=1)
            cell.value = ""
            cell.border = thin_border
            cell.alignment = data_alignment_center
            
            # Column B: Learning Outcome
            cell = ws.cell(row=current_row, column=2)
            cell.value = outcome.get("text", "")
            cell.border = thin_border
            cell.alignment = data_alignment_left
            
            # Column C: RBT
            cell = ws.cell(row=current_row, column=3)
            cell.value = outcome.get("bloom_level", "")
            cell.border = thin_border
            cell.alignment = data_alignment_center
            
            # Column D: No. of Hours
            cell = ws.cell(row=current_row, column=4)
            cell.value = outcome.get("hours", 0)
            cell.border = thin_border
            cell.alignment = data_alignment_center
            
            # ===== BLOOM LEVEL COLUMNS =====
            row_item_total = 0
            row_point_total = 0
            
            for bloom_level, bloom_label in zip(BLOOM_LEVELS, BLOOM_LABELS):
                col = col_map[bloom_label.lower()]
                
                # Get item count from TOS matrix
                item_count = normalized_tos.get(bloom_level, {}).get(row_idx, 0)
                
                # No. of Items
                cell = ws.cell(row=current_row, column=col)
                cell.value = item_count if item_count > 0 else ""
                cell.border = thin_border
                cell.alignment = data_alignment_center
                
                # Points
                cell = ws.cell(row=current_row, column=col + 1)
                cell.value = item_count if item_count > 0 else ""
                cell.border = thin_border
                cell.alignment = data_alignment_center
                
                # Item Numbers (empty - teacher fills)
                cell = ws.cell(row=current_row, column=col + 2)
                cell.value = ""
                cell.border = thin_border
                cell.alignment = data_alignment_center
                
                row_item_total += item_count
                row_point_total += item_count
            
            # ===== ROW TOTALS (ITEMS and POINTS columns) =====
            cell = ws.cell(row=current_row, column=col_map["items_col"])
            cell.value = row_item_total if row_item_total > 0 else ""
            cell.border = thin_border
            cell.alignment = data_alignment_center
            
            cell = ws.cell(row=current_row, column=col_map["points_col"])
            cell.value = row_point_total if row_point_total > 0 else ""
            cell.border = thin_border
            cell.alignment = data_alignment_center
            
            row_totals[row_idx] = {"items": row_item_total, "points": row_point_total}
        
        else:
            # EMPTY ROW - still render with borders
            for col_num in range(1, col_map["points_col"] + 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.value = ""
                cell.border = thin_border
                cell.alignment = data_alignment_center
        
        # Set row height  
        ws.row_dimensions[current_row].height = 20
        current_row += 1
    
    return current_row - 1, row_totals


# ======================================================
# SECTION 4: COMPUTE TOTALS & ROUNDINGS
# ======================================================

def compute_column_totals(outcomes, tos_matrix):
    """
    Compute totals per Bloom level and grand totals.
    
    Returns:
        {bloom: {items, points}, grand_items, grand_points}
    """
    
    normalized_tos = {
        k.strip().lower(): v
        for k, v in tos_matrix.items()
    }
    
    bloom_totals = {}
    grand_items = 0
    grand_points = 0
    
    for bloom_level in BLOOM_LEVELS:
        items_sum = 0
        
        for row_idx in range(len(outcomes)):
            item_count = normalized_tos.get(bloom_level, {}).get(row_idx, 0)
            items_sum += item_count
        
        bloom_totals[bloom_level] = {
            "items": items_sum,
            "points": items_sum
        }
        
        grand_items += items_sum
        grand_points += items_sum
    
    return {
        "bloom_totals": bloom_totals,
        "grand_items": grand_items,
        "grand_points": grand_points
    }


# ======================================================
# SECTION 5: RENDER FOOTER ROWS
# ======================================================

def render_footer_rows(ws, col_map, totals, current_row):
    """
    Render ROUNDINGS and TOTAL rows at bottom.
    Always shown in institutional template.
    """
    
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    header_font = Font(bold=True, size=10)
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add blank row before roundings
    current_row += 1
    
    # ===== ROUNDINGS ROW =====
    ws.cell(row=current_row, column=1).value = ""
    ws.cell(row=current_row, column=1).border = thin_border
    
    cell = ws.cell(row=current_row, column=2)
    cell.value = "ROUNDINGS"
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = center_alignment
    
    ws.cell(row=current_row, column=3).value = ""
    ws.cell(row=current_row, column=3).border = thin_border
    
    ws.cell(row=current_row, column=4).value = ""
    ws.cell(row=current_row, column=4).border = thin_border
    
    # Bloom totals for roundings
    for bloom_level, bloom_label in zip(BLOOM_LEVELS, BLOOM_LABELS):
        col = col_map[bloom_label.lower()]
        items = totals["bloom_totals"][bloom_level]["items"]
        
        cell = ws.cell(row=current_row, column=col)
        cell.value = round(items) if items > 0 else ""
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_alignment
        
        cell = ws.cell(row=current_row, column=col + 1)
        cell.value = round(items) if items > 0 else ""
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_alignment
        
        cell = ws.cell(row=current_row, column=col + 2)
        cell.value = ""
        cell.border = thin_border
    
    # Totals columns for roundings
    cell = ws.cell(row=current_row, column=col_map["items_col"])
    cell.value = round(totals["grand_items"]) if totals["grand_items"] > 0 else ""
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = center_alignment
    
    cell = ws.cell(row=current_row, column=col_map["points_col"])
    cell.value = round(totals["grand_points"]) if totals["grand_points"] > 0 else ""
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = center_alignment
    
    ws.row_dimensions[current_row].height = 20
    current_row += 1
    
    # ===== TOTAL ROW =====
    ws.cell(row=current_row, column=1).value = ""
    ws.cell(row=current_row, column=1).border = thin_border
    
    cell = ws.cell(row=current_row, column=2)
    cell.value = "TOTAL"
    cell.font = Font(bold=True, size=11)
    cell.border = thin_border
    cell.alignment = center_alignment
    
    ws.cell(row=current_row, column=3).value = ""
    ws.cell(row=current_row, column=3).border = thin_border
    
    ws.cell(row=current_row, column=4).value = ""
    ws.cell(row=current_row, column=4).border = thin_border
    
    # Bloom totals for final count
    for bloom_level, bloom_label in zip(BLOOM_LEVELS, BLOOM_LABELS):
        col = col_map[bloom_label.lower()]
        items = totals["bloom_totals"][bloom_level]["items"]
        
        cell = ws.cell(row=current_row, column=col)
        cell.value = items if items > 0 else ""
        cell.font = Font(bold=True, size=11)
        cell.border = thin_border
        cell.alignment = center_alignment
        
        cell = ws.cell(row=current_row, column=col + 1)
        cell.value = items if items > 0 else ""
        cell.font = Font(bold=True, size=11)
        cell.border = thin_border
        cell.alignment = center_alignment
        
        cell = ws.cell(row=current_row, column=col + 2)
        cell.value = ""
        cell.border = thin_border
    
    # Grand totals
    cell = ws.cell(row=current_row, column=col_map["items_col"])
    cell.value = totals["grand_items"] if totals["grand_items"] > 0 else ""
    cell.font = Font(bold=True, size=11)
    cell.border = thin_border
    cell.alignment = center_alignment
    
    cell = ws.cell(row=current_row, column=col_map["points_col"])
    cell.value = totals["grand_points"] if totals["grand_points"] > 0 else ""
    cell.font = Font(bold=True, size=11)
    cell.border = thin_border
    cell.alignment = center_alignment
    
    ws.row_dimensions[current_row].height = 22


# ======================================================
# SECTION 6: APPLY FORMATTING
# ======================================================

def apply_formatting(ws):
    """
    Apply fixed widths and heights for institutional appearance.
    """
    
    # ===== SET COLUMN WIDTHS (FIXED) =====
    ws.column_dimensions['A'].width = 20  # Course Content
    ws.column_dimensions['B'].width = 35  # Learning Outcomes
    ws.column_dimensions['C'].width = 12  # RBT
    ws.column_dimensions['D'].width = 12  # Hours
    
    # Bloom columns (E onwards): 10 units each
    for col_num in range(5, 25):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 10
    
    # ===== SET PRINT AREA =====
    ws.print_area = 'A1:U50'
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5


# ======================================================
# SECTION 7: MAIN EXPORT FUNCTION
# ======================================================

def export_tos_fixed_template(meta, outcomes, tos_matrix, total_items, total_points):
    """
    Master function: Compose all sections into institutional TOS document.
    
    Matches exact structure and formatting from institutional template image.
    
    KEY FIX: Course Content and Learning Outcomes are now separate, distinct 
    columns (not merged), maintaining institutional format exactly.
    
    Args:
        meta: Metadata dictionary
        outcomes: List of outcome dictionaries
        tos_matrix: TOS matrix {bloom: {idx: count}}
        total_items: Total items
        total_points: Total points
    
    Returns:
        BytesIO: Excel file
    """
    
    wb = Workbook()
    ws = wb.active
    ws.title = "TOS"
    
    # Enrich metadata
    meta_enriched = meta.copy()
    meta_enriched["total_items"] = total_items
    meta_enriched["total_points"] = total_points
    
    # Step 1: Render header
    render_institutional_header(ws, meta_enriched)
    
    # Step 2: Create table headers (with separate Content/Outcomes columns)
    col_map = create_table_header_row(ws, start_row=7)
    
    # Step 3: Inject data rows
    last_row, row_totals = inject_data_rows(ws, col_map, outcomes, tos_matrix, start_row=9)
    
    # Step 4: Compute totals
    totals = compute_column_totals(outcomes, tos_matrix)
    
    # Step 5: Render footer rows
    render_footer_rows(ws, col_map, totals, last_row)
    
    # Step 6: Apply formatting
    apply_formatting(ws)
    
    # Step 7: Output
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output
